"""Zaxira nusxa: bitta ro'yxatni Excel (.xlsx) yoki PDF jadval qilib yuklab olish va tiklash.

«💾 Zaxira» bosilganda ro'yxatlar jadvali (№ | Nomi | Summa) chiqadi — foydalanuvchi
raqamini bosib bittasini tanlaydi, so'ng bot formatni so'raydi: 📊 Excel yoki 📄 PDF.
Excel zaxira keyinchalik «📥 Tiklash» orqali qaytarib yuklanadi (eski .json ham);
PDF — chop etish va ulashish uchun.
"""
import html
import io
import json
import logging
import re
from pathlib import Path

from aiogram import Router, F
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

import database as db
from config import MAX_ITEMS
from keyboards import export_format_kb, export_pick_kb, main_menu
from locales import t
from states import Backup
from utils.text import fmt_amount, truncate

router = Router()
logger = logging.getLogger(__name__)

MAX_IMPORT_BYTES = 1_000_000  # 1 MB — zaxira fayli uchun yetarli
MAX_LISTS = 50
_NAME_W = 18  # jadvaldagi «Nomi» ustuni kengligi (telefon ekraniga sig'ishi uchun)
_FONT_DIR = Path(__file__).resolve().parent.parent / "assets" / "fonts"  # PDF shriftlari


# ---------- export ----------

def _lists_table(lang, lists):
    """Ro'yxatlar jadvali (№ | Nomi | Summa) — monospace, <pre> ichida."""
    head = (t(lang, "export_col_num"), t(lang, "export_col_name"), t(lang, "export_col_sum"))
    rows = [
        (str(n), truncate(lst["name"], _NAME_W), fmt_amount(lst["sum"]) if lst["sum"] else "—")
        for n, lst in enumerate(lists, 1)
    ]
    widths = [max(len(head[i]), *(len(r[i]) for r in rows)) for i in range(3)]
    lines = [f"{head[0]:<{widths[0]}}  {head[1]:<{widths[1]}}  {head[2]:>{widths[2]}}"]
    lines += [f"{r[0]:<{widths[0]}}  {r[1]:<{widths[1]}}  {r[2]:>{widths[2]}}" for r in rows]
    return "<pre>" + html.escape("\n".join(lines)) + "</pre>"


@router.callback_query(F.data == "set:export")
async def cb_export(call: CallbackQuery):
    """Butun bazani emas — jadval ko'rsatib, qaysi ro'yxatni zaxiralashni so'raydi."""
    lang = await db.get_lang(call.from_user.id)
    lists = (await db.get_user_lists(call.from_user.id))[:MAX_LISTS]
    if not lists:
        await call.message.answer(t(lang, "export_empty"))
        await call.answer()
        return
    await call.message.answer(
        t(lang, "export_pick_title") + "\n\n" + _lists_table(lang, lists),
        reply_markup=export_pick_kb(lists),
        parse_mode="HTML",
    )
    await call.answer()


def _safe_filename(name, ext):
    """Ro'yxat nomidan fayl nomi yasaydi (taqiqlangan belgilarsiz)."""
    clean = re.sub(r"[^\w \-]", "", name, flags=re.UNICODE).strip()
    return f"{clean[:40] or 'bozorlik'}.{ext}"


def _build_xlsx(lang, list_name, items):
    """Bitta ro'yxatdan .xlsx jadval yasaydi: № | Nomi | Summa + Jami qatori."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Bozorlik"
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14

    ws.merge_cells("A1:C1")
    title = ws.cell(row=1, column=1, value=list_name)
    title.font = Font(bold=True, size=13)

    heads = ("export_col_num", "export_col_name", "export_col_sum")
    for col, key in enumerate(heads, 1):
        ws.cell(row=2, column=col, value=t(lang, key)).font = Font(bold=True)

    total = 0
    row = 3
    for n, item in enumerate(items, 1):
        ws.cell(row=row, column=1, value=n)
        ws.cell(row=row, column=2, value=item["name"])
        if item["price"]:
            price = ws.cell(row=row, column=3, value=item["price"])
            price.number_format = "#,##0"
            total += item["price"]
        row += 1

    ws.cell(row=row, column=2, value=t(lang, "export_total")).font = Font(bold=True)
    cell = ws.cell(row=row, column=3, value=total)
    cell.font = Font(bold=True)
    cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_pdf(lang, list_name, items):
    """Bitta ro'yxatdan PDF jadval yasaydi: № | Nomi | Summa + Jami qatori."""
    pdf = FPDF()
    pdf.add_font("DejaVu", "", str(_FONT_DIR / "DejaVuSans.ttf"))
    pdf.add_font("DejaVu", "B", str(_FONT_DIR / "DejaVuSans-Bold.ttf"))
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("DejaVu", "B", 14)
    pdf.cell(0, 10, list_name)
    pdf.ln(14)

    pdf.set_font("DejaVu", "", 11)
    bold = FontFace(emphasis="BOLD")
    total = 0
    with pdf.table(
        col_widths=(10, 60, 25),
        text_align=("CENTER", "LEFT", "RIGHT"),
        headings_style=FontFace(emphasis="BOLD", fill_color=(235, 235, 235)),
        line_height=8,
    ) as table:
        head = table.row()
        for key in ("export_col_num", "export_col_name", "export_col_sum"):
            head.cell(t(lang, key))
        for n, item in enumerate(items, 1):
            row = table.row()
            row.cell(str(n))
            row.cell(item["name"])
            row.cell(fmt_amount(item["price"]) if item["price"] else "")
            total += item["price"] or 0
        foot = table.row()
        foot.cell("", style=bold)
        foot.cell(t(lang, "export_total"), style=bold)
        foot.cell(fmt_amount(total), style=bold)
    return bytes(pdf.output())


async def _member_list(call, lang, list_id):
    """Ro'yxatni qaytaradi; topilmasa yoki begona bo'lsa — ogohlantirib, None."""
    lst = await db.get_list(list_id)
    if not lst:
        await call.answer(t(lang, "list_gone"), show_alert=True)
        return None
    if not await db.is_member(list_id, call.from_user.id):
        await call.answer(t(lang, "not_member"), show_alert=True)
        return None
    return lst


@router.callback_query(F.data.startswith("exp:"))
async def cb_export_list(call: CallbackQuery):
    """Ro'yxat tanlangach — qaysi formatda yuborishni so'raydi."""
    lang = await db.get_lang(call.from_user.id)
    list_id = int(call.data.split(":")[1])
    lst = await _member_list(call, lang, list_id)
    if not lst:
        return
    await call.message.answer(
        t(lang, "export_format_prompt").format(name=lst["name"]),
        reply_markup=export_format_kb(lang, list_id),
    )
    await call.answer()


@router.callback_query(F.data.startswith("expf:"))
async def cb_export_file(call: CallbackQuery):
    """Tanlangan ro'yxatni tanlangan formatda (.xlsx yoki .pdf) yuboradi."""
    lang = await db.get_lang(call.from_user.id)
    _, list_id, fmt = call.data.split(":")
    lst = await _member_list(call, lang, int(list_id))
    if not lst:
        return
    items = await db.get_items(lst["id"])
    if fmt == "p":
        data, ext, caption_key = _build_pdf(lang, lst["name"], items), "pdf", "export_caption_pdf"
    else:
        data, ext, caption_key = _build_xlsx(lang, lst["name"], items), "xlsx", "export_caption"
    doc = BufferedInputFile(data, filename=_safe_filename(lst["name"], ext))
    await call.message.answer_document(
        doc, caption=t(lang, caption_key).format(name=lst["name"])
    )
    await call.answer()


# ---------- import ----------

@router.callback_query(F.data == "set:import")
async def cb_import(call: CallbackQuery, state: FSMContext):
    lang = await db.get_lang(call.from_user.id)
    await state.set_state(Backup.file)
    await call.message.answer(t(lang, "import_prompt"))
    await call.answer()


async def _unique_name(user_id, name):
    """Import paytida nom takrorlanmasligi uchun oxiriga son qo'shadi."""
    base = name or "Bozorlik"
    existing = {l["name"] for l in await db.get_user_lists(user_id)}
    name, i = base, 2
    while name in existing:
        name = f"{base} ({i})"
        i += 1
    return name


def _as_price(value):
    """Excel katagidagi qiymatni narxga aylantiradi. Yaroqsiz bo'lsa None."""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    if isinstance(value, bool) or not isinstance(value, int):
        return None
    return value if 0 < value <= 10 ** 12 else None


def _parse_xlsx(data):
    """Bot yaratgan .xlsx dan ro'yxatni o'qiydi.

    1-qator — ro'yxat nomi; keyin № raqami va nomi bor qatorlar — mahsulotlar
    (sarlavha va «Jami» qatorlari o'z-o'zidan tushib qoladi).
    """
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        name, items = "", []
        rows = ws.iter_rows(min_row=1, max_row=MAX_ITEMS + 10, max_col=3, values_only=True)
        for idx, row in enumerate(rows, 1):
            a, b, c = (list(row) + [None] * 3)[:3]
            if idx == 1:
                if isinstance(a, str):
                    name = a.strip()
                continue
            if isinstance(a, bool) or not isinstance(a, (int, float)):
                continue
            if not isinstance(b, str) or not b.strip():
                continue
            items.append({"name": b.strip(), "price": _as_price(c), "bought": False})
    finally:
        wb.close()
    return {"name": name, "items": items}


@router.message(Backup.file, F.document)
async def import_file(message: Message, state: FSMContext):
    lang = await db.get_lang(message.from_user.id)
    user_id = message.from_user.id

    if (message.document.file_size or 0) > MAX_IMPORT_BYTES:
        await message.answer(t(lang, "import_bad_file"))
        return

    try:
        buf = await message.bot.download(message.document)
        raw = buf.read()
        if raw[:4] == b"PK\x03\x04":  # .xlsx — zip fayl
            entries = [_parse_xlsx(raw)]
        else:  # eski .json zaxiralar
            payload = json.loads(raw.decode("utf-8"))
            entries = payload["lists"]
            assert isinstance(entries, list)
    except Exception:  # noqa: BLE001 — buzilgan fayl har xil xato berishi mumkin
        await state.clear()
        await message.answer(t(lang, "import_bad_file"), reply_markup=main_menu(lang))
        return

    added = skipped = 0
    for entry in entries[:MAX_LISTS]:
        items = entry.get("items") if isinstance(entry, dict) else None
        if not isinstance(items, list):
            skipped += 1
            continue
        name = await _unique_name(user_id, str(entry.get("name") or "").strip()[:64])
        list_id = await db.create_list(user_id, name)
        for item in items[:MAX_ITEMS]:
            if not isinstance(item, dict):
                continue
            item_name = str(item.get("name") or "").strip()[:64]
            if not item_name:
                continue
            price = item.get("price")
            if not (isinstance(price, int) and 0 < price <= 10 ** 12):
                price = None
            item_id = await db.add_item(list_id, item_name, price)
            if item.get("bought"):
                await db.set_bought(item_id, user_id)
        added += 1
    skipped += max(0, len(entries) - MAX_LISTS)

    await state.clear()
    await message.answer(
        t(lang, "import_done").format(added=added, skipped=skipped),
        reply_markup=main_menu(lang),
    )


@router.message(Backup.file)
async def import_not_a_file(message: Message):
    """Import kutilayotganda hujjat o'rniga matn kelsa — eslatib qo'yamiz."""
    lang = await db.get_lang(message.from_user.id)
    await message.answer(t(lang, "import_prompt"))
